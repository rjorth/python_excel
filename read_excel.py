import openpyxl 


wb = openpyxl.load_workbook('converted_output.xlsx')
#the source is the original CSV for the website
source = wb.get_sheet_by_name('SOURCE')
#this is a list of every sheet created by the pdf 
list_sheets = wb.get_sheet_names()
#remove the source from the list, and useless sheet at the end
list_sheets.remove(list_sheets[0])
list_sheets.remove(list_sheets[-1])

#iterate sheets and search for productnumber/item number
for sheet in list_sheets:
	page = wb.get_sheet_by_name(sheet)
	for i in range(1, 193): #rows in the csv 
		productnumber = source.cell(row=i, column=1).value

		for x in range(1,50):
			check_cell = (page.cell(row=x,column=1).value)

			if check_cell == productnumber:
				print('TRUE\n')
				old_cell_value = source.cell(row=i, column=5).value
				new_cell_value = page.cell(row=x, column=4).value
				print(old_cell_value)
				print(new_cell_value)

				source.cell(row=i,column=5).value = new_cell_value
				print(source.cell(row=i,column=5).value)

			else: pass
wb.save('converted_output.xlsx')



# sheet3 = wb.get_sheet_by_name('Page 3')

# if sheet3['A10'].value == source['A73'].value:
# 	print('TRUE')
# else:
# 	print(sheet3['A10'].value, '\n', source['A73'].value)


